"""Document loading module for the RAG system.

This module provides functionality for loading documents based on their MIME type,
extracting text and metadata, and preparing them for processing.
"""

import asyncio
import logging
from pathlib import Path
from typing import Any

from langchain_community.document_loaders import (
    CSVLoader,
    Docx2txtLoader,
    PDFMinerLoader,
    TextLoader,
    UnstructuredHTMLLoader,
    UnstructuredMarkdownLoader,
    UnstructuredPowerPointLoader,
)
from langchain_core.documents import Document
from openpyxl import load_workbook

from rag.storage.filesystem import FilesystemManager
from rag.utils.exceptions import (
    DocumentLoadingError,
    LoaderInitializationError,
    UnsupportedFileError,
)
from rag.utils.logging_utils import log_message

from .metadata_extractor import DocumentMetadataExtractor

logger = logging.getLogger(__name__)


class SimpleExcelLoader:
    """Minimal loader for Excel ``.xlsx`` files."""

    def __init__(self, file_path: str) -> None:
        """Create loader for Excel files."""
        self.file_path = file_path

    def load(self) -> list[Document]:
        """Load workbook contents as a single document."""
        wb = load_workbook(self.file_path, read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                values = [str(cell) for cell in row if cell is not None]
                if values:
                    lines.append(" ".join(values))
        return [Document(page_content="\n".join(lines))]


class DocumentLoader:
    """Loads documents based on their MIME type.

    This class provides functionality for loading documents from different file types,
    extracting text and metadata, and preparing them for processing.
    """

    def __init__(
        self,
        filesystem_manager: FilesystemManager,
        log_callback: Any | None = None,
    ) -> None:
        """Initialize the document loader.

        Args:
            filesystem_manager: Filesystem manager for file operations
            log_callback: Optional callback for logging

        """
        self.filesystem_manager = filesystem_manager
        self.log_callback = log_callback
        self.metadata_extractor = DocumentMetadataExtractor()
        self.last_loader_name: str | None = None

    def _log(self, level: str, message: str) -> None:
        """Log a message.

        Args:
            level: Log level (INFO, WARNING, ERROR, etc.)
            message: The log message

        """
        log_message(level, message, "DocumentLoader", self.log_callback)

    def get_loader_for_file(self, file_path: Path | str) -> Any:
        """Get the appropriate loader for a file based on its MIME type.

        Args:
            file_path: Path to the file

        Returns:
            Langchain document loader instance

        Raises:
            LoaderInitializationError: If the loader could not be initialized
            UnsupportedFileError: If the file type is not supported

        """
        file_path = Path(file_path)
        mime_type = self.filesystem_manager.get_file_type(file_path)

        # Map MIME types to loader classes
        loader_map = {
            "text/plain": TextLoader,
            "text/csv": CSVLoader,
            "text/markdown": UnstructuredMarkdownLoader,
            "text/html": UnstructuredHTMLLoader,
            "application/pdf": PDFMinerLoader,
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document": Docx2txtLoader,
            "application/msword": Docx2txtLoader,
            "application/vnd.openxmlformats-officedocument.presentationml.presentation": UnstructuredPowerPointLoader,
            "application/vnd.ms-powerpoint": UnstructuredPowerPointLoader,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": SimpleExcelLoader,
        }

        # Special handling for text files with different extensions
        if mime_type == "text/plain":
            suffix = file_path.suffix.lower()
            if suffix == ".csv":
                return CSVLoader(str(file_path))
            if suffix in [".md", ".markdown"]:
                return UnstructuredMarkdownLoader(str(file_path))
            if suffix in [".json", ".yml", ".yaml"]:
                return TextLoader(str(file_path))

        # Get loader class for the MIME type
        loader_class = loader_map.get(mime_type)
        if not loader_class:
            self._log(
                "WARNING",
                f"Unsupported file type: {mime_type}, attempting TextLoader",
            )
            loader_class = TextLoader

        # Initialize the loader
        try:
            return loader_class(str(file_path))
        except Exception as e:
            self._log("ERROR", f"Failed to initialize loader for {file_path}: {e}")
            raise LoaderInitializationError(file_path, str(e)) from e

    def load_document(self, file_path: Path | str) -> list[Document]:
        """Load a document from a file.

        Args:
            file_path: Path to the file

        Returns:
            List of langchain Document objects

        Raises:
            UnsupportedFileError: If the file is not supported or doesn't exist
            DocumentLoadingError: If the file could not be loaded

        """
        file_path = Path(file_path)

        # Check if file exists and is supported
        if not self.filesystem_manager.is_supported_file(file_path):
            self._log("ERROR", f"Unsupported or non-existent file: {file_path}")
            raise UnsupportedFileError(file_path)

        # Get loader for the file
        try:
            loader = self.get_loader_for_file(file_path)
            self.last_loader_name = loader.__class__.__name__
            self._log("DEBUG", f"Loading document: {file_path}")

            # Handle both sync and async loaders
            if hasattr(loader, "aload"):
                # Create a new event loop if we're not in one
                try:
                    loop = asyncio.get_event_loop()
                except RuntimeError:
                    loop = asyncio.new_event_loop()
                    asyncio.set_event_loop(loop)

                # Run the async loader
                docs = loop.run_until_complete(loader.aload())
            else:
                # Use the sync loader
                docs = loader.load()

            # Add file metadata to each document
            self._enhance_document_metadata(docs, file_path)

            # Get MIME type of the file
            mime_type = self.filesystem_manager.get_file_type(file_path)

            # Extract and add document-specific metadata
            self._log("DEBUG", f"Extracting metadata from document: {file_path}")
            docs = self.metadata_extractor.enhance_documents(docs, mime_type)

            self._log("DEBUG", f"Loaded {len(docs)} document(s) from {file_path}")
        except Exception as e:
            self._log("ERROR", f"Failed to load document {file_path}: {e}")
            raise DocumentLoadingError(file_path, str(e)) from e
        else:
            return docs

    def _enhance_document_metadata(
        self,
        docs: list[Document],
        file_path: Path | str,
    ) -> None:
        """Enhance document metadata with file information.

        Args:
            docs: List of documents to enhance
            file_path: Path to the source file
        """
        file_path = Path(file_path)
        for doc in docs:
            # Add file metadata if not already present
            if "source" not in doc.metadata:
                doc.metadata["source"] = str(file_path)
            if "file_name" not in doc.metadata:
                doc.metadata["file_name"] = file_path.name
            if "file_type" not in doc.metadata:
                doc.metadata["file_type"] = self.filesystem_manager.get_file_type(
                    file_path
                )
            if "loader_name" not in doc.metadata:
                doc.metadata["loader_name"] = self.last_loader_name
